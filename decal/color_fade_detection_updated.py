import cv2
import numpy as np
import openpyxl

# Load the Excel workbook and select the active worksheet
workbook = openpyxl.load_workbook(r'C:\Users\M_Thiruveedula\Downloads\Code Tester\decal\example.xlsx')
worksheet = workbook.active

# Loop through each row in the worksheet
for row in worksheet.iter_rows(min_row=2, values_only=True):
    # Read the original and current image file paths from the current row
    original_image_path = row[0]
    print(original_image_path)
    current_image_path = row[1]
    print(current_image_path)
    
    # Load the original and current images
    original_image = cv2.imread(original_image_path)
    current_image = cv2.imread(current_image_path)

    # Set a threshold for color difference
    threshold = 30

    # Calculate the absolute difference between the images
    difference = cv2.absdiff(original_image, current_image)

    # Split the difference image into separate color channels
    b, g, r = cv2.split(difference)

    # Check if any pixel in any channel has a value greater than the threshold
    if cv2.countNonZero(np.array(b > threshold, dtype=np.uint8)) or cv2.countNonZero(np.array(g > threshold, dtype=np.uint8)) or cv2.countNonZero(np.array(r > threshold, dtype=np.uint8)):
        outcome = "Significant color difference detected!"
    else:
        outcome = "No significant color difference detected."

    # Write the outcome to the current row in the Excel worksheet
    worksheet.cell(row=row[2], column=3, value=outcome)

# Save the modified workbook
workbook.save(r'C:\Users\M_Thiruveedula\Downloads\Code Tester\decal\example.xlsx')
